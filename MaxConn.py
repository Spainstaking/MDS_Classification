import matplotlib.pyplot as plt
import networkx as nx
import openpyxl

'''求解极大连通子图'''

'''
    function:
        读入Excel文档（默认带权重）
    input:
        file:输入表格(xlsx格式)：第一列为源项目，第二列为引用项目，第三列为权重(可选)
    output:
        nodes:节点(项目)
        edges:边(依赖关系)
'''


def readExcel(file):  # cls : 表示没用被实例化的类本身
    nodes = []
    edges = []
    # 输入的表格必须整行整列（即m*n矩阵中的每个单元格不能为空）
    wb = openpyxl.load_workbook(file)
    mySheet = wb['Sheet1']
    m_rows = mySheet.max_row
    n_columns = mySheet.max_column
    for i in range(1, m_rows + 1):
        mylist = []
        for j in range(1, n_columns + 1):
            key = mySheet.cell(i, j).value
            mylist.append(key)  # 该列 添加至 mylist
            # key not in nodes
            if (j < 3) and (not judgeIsIn(nodes, key)):  # 找到边的两个节点
                nodes.append(key)
        tup = tuple(mylist)
        edges.append(tup)
    # wb.save(file)
    return nodes, edges


def readTXT(file):  # cls : 表示没用被实例化的类本身
    nodes = []
    edges = []
    for line in open(file, "r"):
        mylist = []
        node1, node2 = line.strip('\n').split(' ')

        if not judgeIsIn(nodes, node1):
            nodes.append(node1)
        if not judgeIsIn(nodes, node2):
            nodes.append(node2)

        mylist.append(node1)
        mylist.append(node2)
        tup = tuple(mylist)
        edges.append(tup)
    return nodes, edges


'''
def itsays_init():

  

    dict = {}

    mylist1 = []
    mylist2 = []
    for edge in edges:
        node1 = edge[0]
        node2 = edge[1]
        mylist1.append(node1)  #左边列表
        mylist2.append(node2)  #右边列表
    for key in nodes:
        num1 = mylist1.count(key)  # 返回元素在列表中出现的次数
        num2 = mylist2.count(key)
        num = num1 + num2
        dict[key] = num
    nodesdegree = dict
'''

'''
function:
    读入节点和边
input:
    nodes:节点(项目)
    edges:边(依赖关系)
    weight_b：是否带权重(默认为true)
'''


def addNodesEdges(G, weight_b, nodes, edges):
    # 定义graph
    G.add_nodes_from(nodes)
    if weight_b:
        G.add_weighted_edges_from(edges)
    else:
        G.add_edges_from(edges)
    return


'''
function:
    将极大连通子图写入Excel文档（默认带权重）
input:
    mylist：连通子图
output:
    file:输出表格(xlsx格式)：第一列为源项目，第二列为引用项目，第三列为权重(可选)
'''


def writeGraph(G, file):
    # 写文件
    wb = openpyxl.Workbook()  # 创建新的excel文件，一个工作簿(workbook)在创建的时候同时至少也新建了一张工作表(worksheet)
    # wb.save(file)
    wb.create_sheet(index=0, title="Sheet1")  # 可通过index控制创建的表的位置
    sheet_names = wb.sheetnames  # 获取所有表名
    sheet1 = wb[sheet_names[0]]  # 打开第一个 sheet 工作表

    i = 1
    print("写入极大连通子图")
    for sublist in nx.connected_components(G):
        j = 1
        for key in sublist:
            sheet1.cell(row=i, column=j, value=key)
            j += 1
        i += 1
        # print("写入极大连通子图数：", i)
    wb.save(file)
    return


'''
计算极大连通子图中   不同图的数量及节点个数
统计极大连通子图
    input:
        mylist：连通子图
'''


def count(G, num):
    dict = {}
    for tmp_list in nx.connected_components(G):
        length = len(tmp_list)  # 当前连通图中   的  节点个数
        ret = dict.get(length)  # 拥有相同节点个数  的图的个数
        if ret is None:
            dict[length] = 1
        else:
            dict[length] = ret + 1
    print(dict[num])
    # items = sorted(dict.items(), key=lambda item: item[0])
    # for key in items:
    #     print(key)  # 如(4,1)指的是 节点数为4的子图个数为1
    return






'''
过滤社区中节点数小于threshold的极大连通子图
input:
    mylist:极大连通子图 nx.connected_components(G)
    threshold:节点阈值
output:
    节点列表
'''


def filtercommunity(G, threshold):
    dataset = []
    for tmplist in nx.connected_components(G):
        if len(tmplist) >= threshold:
            for key in tmplist:
                if judgeIsIn(dataset, key):
                    continue
                else:  # 单元素点不在dataset中
                    dataset.append(key)
    # self.dataset = dataset.copy()  # 只拷贝父对象，不拷贝子对象
    print(dataset)
    return


'''
求最大的 极大连通子图
'''


def maxSubGraph(G):
    maxsubgraph = []
    max = 0
    for tmplist in nx.connected_components(G):
        if max < len(tmplist):
            max = len(tmplist)
            maxsubgraph.clear()
            # maxsubgraph = tmplist.copy()
            maxsubgraph = G.subgraph(tmplist)

    print('最大连通子图的节点数：', max)
    print(maxsubgraph)
    return maxsubgraph, max


'''
function:
    显示图形
input:
    mylist：连通子图
'''


def show(G):
    plt.subplot(111)
    # print(len(max(nx.connected_component(G))))
    nx.draw_networkx(G, with_labels=True)  # 有框框
    '''子图
    # color = ['y', 'g']
    # subplot = [323, 324]
    # # 打印连通子图
    
    # for c in nx.connected_components(G):
    #     # 得到不连通的子集
    #     nodeSet = G.subgraph(c).nodes()
    #     # 绘制子图
    #     subgraph = G.subgraph(c)
    #     plt.subplot(subplot[0])  # 第二整行
    #     nx.draw_networkx(subgraph, with_labels=True, node_color=color[0])
    # 
    #     color.pop(0)
    #     subplot.pop(0)
    '''
    # plt.subplot(312)

    # nx.draw_networkx(G, with_labels=True, node_color ='y')
    plt.show()
    return


def judgeIsIn(data_all, temp):
    for k in data_all:
        if k == temp:
            return True
    return False


if __name__ == "__main__":

    # nodes, edges = ConnectedSubgraph.readExcel(r"C:\Users\SureY\Desktop\Normal_BRCA_edge.xlsx")
    # nodes, edges = ConnectedSubgraph.readExcel(r"C:\Users\SureY\Desktop\1.xlsx")
    raise Exception('gggg')
    file_names = os.listdir(r'E:\PycharmTest\MDS\PPI_PCCs_edge')

    for fname in file_names:

        file_tumor = os.path.join("PPI_PCCs_edge", fname)
        for fname2 in os.listdir(file_tumor):
            file = os.path.join("PPI_PCCs_edge", fname, fname2)
            file_maxconn = os.path.splitext(os.path.basename(file))[0] + '_maxconn.txt'
            file_maxconn_path = os.path.join("PPI_PCCs_edge_maxconn", fname, file_maxconn)

            nodes, edges = readTXT(file)
            # 定义graph
            G = nx.Graph()
            addNodesEdges(G, False)
            # connectedsubgraph.count(10241)  # 如(4,1)指的是 节点数为4的子图个数为1
            maxsubgraph, max = maxSubGraph(G)
            # connectedsubgraph.filtercommunity(4)  # 连通图>= 4的结点列表
            # connectedsubgraph.show()
            with open(file_maxconn_path, 'a') as fp:
                for i in maxsubgraph.edges:
                    fp.write(i[0] + ' ' + i[1] + '\n')
            print(fname2 + '  over')
